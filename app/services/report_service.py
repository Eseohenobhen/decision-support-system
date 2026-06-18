from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.models.fund_transaction import FundTransaction
from app.models.property import Property
from app.models.property_fund import PropertyFund
from app.models.maintenance_request import MaintenanceRequest
from app.models.user import User, UserRole
from app.services.manager_access import manager_property_access_filter


def _apply_property_manager_filter(stmt, current_user):
    if current_user.role == UserRole.PROPERTY_MANAGER:
        stmt = stmt.where(manager_property_access_filter(current_user.id))
    return stmt


def _apply_property_fund_manager_filter(stmt, current_user):
    if current_user.role == UserRole.PROPERTY_MANAGER:
        stmt = stmt.join(Property).where(manager_property_access_filter(current_user.id))
    return stmt


def _apply_transaction_manager_filter(stmt, current_user):
    if current_user.role == UserRole.PROPERTY_MANAGER:
        stmt = stmt.join(FundTransaction.fund).join(Property).where(manager_property_access_filter(current_user.id))
    return stmt


def _apply_request_manager_filter(stmt, current_user):
    if current_user.role == UserRole.PROPERTY_MANAGER:
        stmt = stmt.join(Property).where(manager_property_access_filter(current_user.id))
    return stmt


def get_property_summary(db: Session, current_user: User) -> list[dict[str, object]]:
    stmt = select(Property).options(joinedload(Property.manager)).order_by(Property.name)
    stmt = _apply_property_manager_filter(stmt, current_user)
    properties = list(db.scalars(stmt))

    return [
        {
            "Property Code": property_.code,
            "Property Name": property_.name,
            "City": property_.city,
            "State": property_.state,
            "Country": property_.country,
            "Manager": property_.manager.full_name if property_.manager else "-",
            "Total Units": property_.total_units,
            "Created At": property_.created_at.strftime("%Y-%m-%d"),
        }
        for property_ in properties
    ]


def get_fund_allocation_report(db: Session, current_user: User) -> list[dict[str, object]]:
    stmt = select(PropertyFund).options(joinedload(PropertyFund.property), joinedload(PropertyFund.allocations)).order_by(PropertyFund.fiscal_year.desc())
    stmt = _apply_property_fund_manager_filter(stmt, current_user)
    funds = list(db.scalars(stmt))

    return [
        {
            "Property Code": fund.property.code if fund.property else "-",
            "Property Name": fund.property.name if fund.property else "-",
            "Fiscal Year": fund.fiscal_year,
            "Annual Budget": float(fund.annual_budget),
            "Current Balance": float(fund.current_balance),
            "Reserved Balance": float(fund.reserved_balance),
            "Allocated Items": len(fund.allocations),
            "Currency": fund.currency,
        }
        for fund in funds
    ]


def get_maintenance_history(db: Session, current_user: User) -> list[dict[str, object]]:
    stmt = select(MaintenanceRequest).options(joinedload(MaintenanceRequest.property)).order_by(MaintenanceRequest.requested_at.desc())
    stmt = _apply_request_manager_filter(stmt, current_user)
    requests = list(db.scalars(stmt))

    return [
        {
            "Request ID": request.id,
            "Property Code": request.property.code if request.property else "-",
            "Title": request.title,
            "Status": request.status.value,
            "Priority": request.priority.value,
            "Priority Score": request.priority_score,
            "Estimated Cost": float(request.estimated_cost),
            "Approved Cost": float(request.approved_cost or 0),
            "Requested At": request.requested_at.strftime("%Y-%m-%d"),
            "Completed At": request.completed_at.strftime("%Y-%m-%d") if request.completed_at else "",
        }
        for request in requests
    ]


def get_monthly_financial_report(db: Session, current_user: User) -> list[dict[str, object]]:
    stmt = select(FundTransaction).options(joinedload(FundTransaction.fund).joinedload(PropertyFund.property))
    stmt = _apply_transaction_manager_filter(stmt, current_user)
    transactions = list(db.scalars(stmt.order_by(FundTransaction.transaction_date.desc())))

    month_groups: dict[str, dict[str, object]] = defaultdict(lambda: {
        "month": "",
        "Credits": 0.0,
        "Debits": 0.0,
        "Reserves": 0.0,
        "Releases": 0.0,
        "Net Change": 0.0,
    })

    for tx in transactions:
        if not tx.transaction_date:
            continue

        key = tx.transaction_date.strftime("%Y-%m")
        group = month_groups[key]
        group["month"] = tx.transaction_date.strftime("%b %Y")
        amount = float(tx.amount)

        if tx.transaction_type.value == "credit":
            group["Credits"] += amount
            group["Net Change"] += amount
        elif tx.transaction_type.value == "debit":
            group["Debits"] += amount
            group["Net Change"] -= amount
        elif tx.transaction_type.value == "reserve":
            group["Reserves"] += amount
            group["Net Change"] -= amount
        elif tx.transaction_type.value == "release":
            group["Releases"] += amount
            group["Net Change"] += amount
        else:
            group["Net Change"] += 0.0

    sorted_months = sorted(month_groups.items(), reverse=True)
    return [group for _, group in sorted_months]


def _create_pdf_report(title: str, headers: list[str], rows: list[dict[str, object]]) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = [Paragraph(title, title_style), Spacer(1, 8), Paragraph(f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}", meta_style), Spacer(1, 16)]

    data = [headers]
    for row in rows:
        data.append([row.get(key, "") for key in headers])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def _create_excel_report(title: str, headers: list[str], rows: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="left")

    sheet.cell(row=2, column=1, value=f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}")
    header_fill = PatternFill(fill_type="solid", start_color="FFe2e8f0", end_color="FFe2e8f0")

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=5):
        for col_index, header in enumerate(headers, start=1):
            cell_value = row.get(header, "")
            if isinstance(cell_value, Decimal):
                cell_value = float(cell_value)
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    for column_index in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(row=4, column=column_index).column_letter].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_report_file(report_name: str, rows: list[dict[str, object]], format: str) -> BytesIO:
    if not rows:
        rows = [{"Message": "No records found"}]
    headers = list(rows[0].keys())
    if format == "pdf":
        return _create_pdf_report(report_name, headers, rows)
    return _create_excel_report(report_name, headers, rows)
